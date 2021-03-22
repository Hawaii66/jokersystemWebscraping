from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
import getpass
from time import sleep

driver = webdriver.Chrome()
driver.get("https://www.jokersystemet.se/")

def Login(driver):
    loginElemenet = driver.find_element_by_xpath("/html/body/div[1]/header/div/div/div[2]/ul/li[9]/a/div/span")
    loginElemenet.click()
    emailElement = driver.find_element_by_xpath("/html/body/div[1]/main/form/input[1]")
    emailElement.send_keys("") #EMAIL
    passwordElement = driver.find_element_by_xpath("/html/body/div[1]/main/form/input[2]")
    passwordElement.send_keys("") #PASSWORD
    passwordElement.submit()


def findToSpelTyp(driver):
    filename = "joker_out.xlsx"
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    sheet["N11"] = "V75"

    return sheet["N11"].value


def GoToV(driver, number):
    print(number)
    if(number == "V86"):
        v86Element = driver.find_element_by_xpath("/html/body/div[1]/main/div[1]/div[2]/div/div[1]/section[3]/section/ul/li[4]/button/div")
        v86Element.click()
        sleep(1)
        goElement = driver.find_element_by_xpath("/html/body/div[1]/div[3]/div/ul/li[1]/a")
        goElement.click()
        return
    if(number == "V75"):
        v75Element = driver.find_element_by_xpath("/html/body/div[1]/main/div[1]/div[2]/div/div[1]/section[3]/section/ul/li[5]/button/div")
        v75Element.click()
        sleep(1)
        goElement = driver.find_element_by_xpath("/html/body/div[1]/div[3]/div/ul/li[1]/a")
        goElement.click()
    if(number == "V65"):
        v65Element = driver.find_element_by_xpath("/html/body/div[1]/main/div[1]/div[2]/div/div[1]/section[3]/section/ul/li[2]/button/div")
        v65Element.click()
        sleep(1)
        goElement =driver.find_element_by_xpath("/html/body/div[1]/div[3]/div/ul/li[1]/a")
        goElement.click()
    if(number == "V64"):
        v64Element = driver.find_element_by_xpath("/html/body/div[1]/main/div[1]/div[2]/div/div[1]/section[3]/section/ul/li[3]/button/div")
        v64Element.click()
        sleep(1)
        goElement = driver.find_element_by_xpath("/html/body/div[1]/div[3]/div/ul/li[1]/a")
        goElement.click()
    if(number == "V4"):
        v4Element = driver.find_element_by_xpath("/html/body/div[1]/main/div[1]/div[2]/div/div[1]/section[3]/section/ul/li[1]/button/div")
        v4Element.click()
        sleep(1)
        goElement = driver.find_element_by_xpath("/html/body/div[1]/div[3]/div/ul/li[1]/a")
        goElement.click()


    #if(number == "V5"):
    #    pass
    #if(number == "V4"):
    #    pass
    #if(number == "V3"):
    #    pass


def switchToTrending(driver):
    swithElement = driver.find_element_by_xpath("/html/body/div[1]/main/div/div[1]/button")
    swithElement.click()
    trendButton = driver.find_element_by_xpath("/html/body/div[4]/div[8]")
    trendButton.click()
    senaste30Element = driver.find_element_by_xpath("/html/body/div[1]/main/section/div[1]/div/div[2]/button[1]")
    senaste30Element.click()


def findAllTrending(driver, multPercent):

    rowParent = driver.find_element_by_class_name("rows")

    rows = rowParent.find_elements_by_class_name("row")

    print(rows)

    allHorses = []

    for row in rows:
        print(row)
        print("------")

        cellItems = row.find_elements_by_class_name("item")

        print(cellItems)

        #horseInfo = {
        #    "horseNumber":0,
        #    "percent":0,
        #    "trending":0
        #}
        horses = []

        for cell in cellItems:
            number = cell.find_element_by_class_name("number").text
            percent = cell.find_element_by_class_name("info-1").text
            trending = cell.find_element_by_class_name("trending").text
            odds = cell.find_element_by_class_name("info-2").text
            print("Horse Number: " + number)
            print("Horse Percent: " + percent)
            print("Horse Trending: " + trending)

            percent = percent[:-1]

            if(trending == "-"):
                jTrending = 0
            else:
                jTrending = trending

            #print(trending)
            #print(jTrending)
            #print(multPercent)

            if(odds == ""):
                odds = 999

            horseInfo = {
                "horseNumber":number,
                "percent":percent,
                "trending":trending,
                "odds":odds,
                "jStrP":(float(jTrending) * multPercent) + float(percent)
            }
            horses.append(horseInfo)

        print(horses)
        allHorses.append(horses)


    return allHorses


def findMultPercent(driver):
    filename = "joker_out.xlsx"
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    sheet["M5"] = "Typ av spel"
    sheet["M6"] = "V86"
    sheet["M7"] = "V75"
    sheet["M8"] = "V65"
    sheet["M9"] = "V64"
    sheet["N6"] = 1.5
    sheet["N7"] = 0.75
    sheet["N8"] = 3.4
    sheet["N9"] = 1.2
    sheet["M11"] = "Kör"
    sheet["N11"] = "V75"

    currentLopp = sheet["N11"].value
    multPercent = 0
    if(currentLopp == "V86"):
        multPercent = sheet["N6"].value
    elif(currentLopp == "V75"):
        multPercent = sheet["N7"].value
    elif (currentLopp == "V65"):
        multPercent = sheet["N8"].value
    elif (currentLopp == "V64"):
        multPercent = sheet["N9"].value

    print(multPercent)

    workbook.save(filename=filename)

    return multPercent


def sortHorses(horses):
    newList = []
    for j in range(len(horses)):
        currentLowest = {}
        for i in range(len(horses)):
            horse = horses[i]
            print(i)
            print(currentLowest)
            if(currentLowest == {}):
                currentLowest = horse
            else:
                if(currentLowest["jStrP"] < horse["jStrP"]):
                    currentLowest = horse
        horses.remove(currentLowest)
        newList.append(currentLowest)
    return newList


def writeToExcell(allHorses):
    filename = "joker_out.xlsx"

    workbook = Workbook()
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    columns = ["B", "C", "D", "E", "F", "G"]

    print("TEST")

    print(allHorses)

    offset = 2

    #Clear previous Run

    for a in sheet["A1":"G150"]:
        print("HKAJSHgkjASHGJKASHGKJAHGKAS1111111111111")
        for cell in a:
            print("AHGHAJGHJKASHgkASGKAHSKGhKASHGKASg")
            cell.value = None


    for horses in allHorses:
        index = offset
        sheet[columns[0] + str(index)] = "Nr"
        sheet[columns[1] + str(index)] = "Str%"
        sheet[columns[2] + str(index)] = "Trend"
        sheet[columns[3] + str(index)] = "Odds"
        sheet[columns[4] + str(index)] = "JStr%"
        sheet[columns[5] + str(index)] = "OldRank"

        index += 1

        oldHorses = horses.copy()
        print(oldHorses)
        print("------------")
        horses = sortHorses(horses)
        print(oldHorses)
        oldIndex = index
        for horse in horses:
            horseID = horse["horseNumber"]
            horsePercent = horse["percent"]
            horseTrending = horse["trending"]
            horseOdds = horse["odds"]
            #print(horseID)

            if(horseTrending == "-"):
                horseTrending = "0"

            jStrP = horse["jStrP"]

            if(jStrP < 0):
                jStrP = 0

            sheet[columns[0] + str(index)] = int(horseID)
            sheet[columns[1] + str(index)] = float(horsePercent)
            sheet[columns[2] + str(index)] = float(horseTrending)
            sheet[columns[3] + str(index)] = float(horseOdds)
            sheet[columns[4] + str(index)] = float(jStrP)
            index += 1

        index = oldIndex
        print("TSET")
        print(oldHorses)
        for oldHorse in oldHorses:
            horseID = oldHorse["horseNumber"]
            print(horseID)
            sheet[columns[5] + str(index)] = int(horseID)
            index += 1

        offset += 17

    workbook.save(filename=filename)

#toPlay = input("vad ska spelas")
Login(driver)
sleep(2)
GoToV(driver, findToSpelTyp(driver))
sleep(2)
switchToTrending(driver)
sleep(2)
alLHorses = findAllTrending(driver, findMultPercent(driver))
writeToExcell(alLHorses)

#content = driver.find_elements_by_class_name("trending")
#print(content)